"""
cleaner.py — cleans jobs_raw.csv, extracts skills, loads into SQLite + Excel
Run: python cleaner.py
"""
import os, re, sqlite3
import pandas as pd
from datetime import datetime
from config import DB_PATH, CSV_OUTPUT, XLSX_OUTPUT, DATA_DIR, TARGET_LOCATIONS


 # Master skill list — order matters (longer phrases first to avoid partial matches)
SKILLS = [
    # SQL family
    "google bigquery","ms sql server","sql server","postgresql","mysql","sqlite","bigquery","snowflake","redshift","sql",
    # Python ecosystem
    "scikit-learn","matplotlib","seaborn","numpy","pandas","pyspark","python",
    # BI tools
    "power bi","microsoft power bi","looker studio","google data studio","qlik sense","qlikview","tableau","looker","qlik",
    # Excel
    "advanced excel","power query","power pivot","ms excel","excel",
    # Cloud
    "google cloud","aws","azure","gcp",
    # ML/AI
    "machine learning","deep learning","natural language processing","large language model",
    "prompt engineering","generative ai","llm","nlp","ai","ml",
    # ETL/Data Engineering
    "apache spark","apache kafka","apache airflow","dbt","hadoop","kafka","airflow","spark","etl",
    # Other languages
    "r programming","scala","java","r",
    # Databases
    "mongodb","cassandra","hive","presto","databricks",
    # Soft/domain
    "statistics","data visualization","data modelling","data modeling","data warehousing",
    "business intelligence","ab testing","a/b testing",
]

CITY_MAP = {
    "bengaluru":"bangalore","bangaluru":"bangalore","blr":"bangalore",
    "new delhi":"delhi","ncr":"delhi","delhi ncr":"delhi",
    "noida/greater noida":"noida",
    "gurugram":"gurgaon","gurgaon/gurugram":"gurgaon",
    "hyderabad/secunderabad":"hyderabad",
    "navi mumbai":"mumbai","thane":"mumbai",
}


def normalize_city(loc):
    if not loc or str(loc) in ("N/A","nan",""): return "other"
    l = str(loc).lower()
    for k,v in CITY_MAP.items():
        if k in l: return v
    for c in ["bangalore","pune","hyderabad","mumbai","delhi","chennai","gurgaon","noida"]:
        if c in l: return c
    return "other"

def extract_skills(text):
    if not text or str(text) in ("N/A","nan",""): return ""
    t = str(text).lower()
    return ", ".join(sk for sk in SKILLS if re.search(r'\b'+re.escape(sk)+r'\b', t))

def parse_exp(exp):
    if not exp or str(exp) in ("N/A","nan",""): return None, None
    nums = re.findall(r'\d+', str(exp))
    if len(nums)>=2: return int(nums[0]),int(nums[1])
    if len(nums)==1: return int(nums[0]),int(nums[0])
    return None,None

def categorize(title):
    t = str(title).lower()
    if any(x in t for x in ["data engineer","etl","pipeline"]): return "Data Engineer"
    if any(x in t for x in ["business analyst","bi analyst","mis"]): return "Business Analyst"
    if any(x in t for x in ["power bi","tableau","bi developer"]): return "BI Developer"
    if any(x in t for x in ["scientist","machine learning","ai engineer"]): return "Data Scientist/ML"
    return "Data Analyst"

def filter_locations(df):
    if not TARGET_LOCATIONS: return df
    pattern = "|".join(TARGET_LOCATIONS)
    mask = df["Location"].str.contains(pattern, case=False, na=False)
    filtered = df[mask]
    print(f"  Location filter: {len(df)} → {len(filtered)}")
    return filtered

SCHEMA = """
CREATE TABLE IF NOT EXISTS jobs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    job_title TEXT, company TEXT, location TEXT,
    city_normalized TEXT, role_category TEXT,
    experience TEXT, exp_min REAL, exp_max REAL,
    salary TEXT, salary_min_lpa REAL, salary_max_lpa REAL,
    skills_extracted TEXT, skills_raw TEXT, nice_to_have TEXT,
    is_fresher_role INTEGER, posted_date TEXT,
    scrape_date TEXT, scrape_week TEXT,
    job_url TEXT, search_query TEXT, search_city TEXT, platform TEXT,
    UNIQUE(job_url, scrape_date)
);
CREATE TABLE IF NOT EXISTS weekly_skill_trends (
    week TEXT, skill TEXT, mention_count INTEGER, PRIMARY KEY(week,skill));
CREATE TABLE IF NOT EXISTS weekly_city_demand (
    week TEXT, city TEXT, job_count INTEGER,
    avg_salary_min REAL, fresher_pct REAL, PRIMARY KEY(week,city));
CREATE TABLE IF NOT EXISTS weekly_role_demand (
    week TEXT, role_category TEXT, job_count INTEGER, PRIMARY KEY(week,role_category));
CREATE TABLE IF NOT EXISTS weekly_company_demand (
    week TEXT, company TEXT, job_count INTEGER, PRIMARY KEY(week,company));
"""

def aggregate(conn, week):
    cur = conn.cursor()
    rows = cur.execute("SELECT skills_extracted FROM jobs WHERE scrape_week=? AND skills_extracted!=''", (week,)).fetchall()
    counts = {}
    for (s,) in rows:
        for sk in s.split(","):
            sk=sk.strip()
            if sk: counts[sk]=counts.get(sk,0)+1
    cur.executemany("INSERT OR REPLACE INTO weekly_skill_trends VALUES(?,?,?)", [(week,sk,cnt) for sk,cnt in counts.items()])
    rows = cur.execute("""SELECT city_normalized,COUNT(*),AVG(salary_min_lpa),
        ROUND(100.0*SUM(is_fresher_role)/COUNT(*),1)
        FROM jobs WHERE scrape_week=? AND city_normalized!='other' GROUP BY city_normalized""",(week,)).fetchall()
    cur.executemany("INSERT OR REPLACE INTO weekly_city_demand VALUES(?,?,?,?,?)", [(week,r[0],r[1],r[2],r[3]) for r in rows])
    rows = cur.execute("SELECT role_category,COUNT(*) FROM jobs WHERE scrape_week=? GROUP BY role_category",(week,)).fetchall()
    cur.executemany("INSERT OR REPLACE INTO weekly_role_demand VALUES(?,?,?)", [(week,r[0],r[1]) for r in rows])
    rows = cur.execute("SELECT company,COUNT(*) FROM jobs WHERE scrape_week=? AND company!='' GROUP BY company ORDER BY COUNT(*) DESC LIMIT 50",(week,)).fetchall()
    cur.executemany("INSERT OR REPLACE INTO weekly_company_demand VALUES(?,?,?)", [(week,r[0],r[1]) for r in rows])
    conn.commit()

def load_db(df):
    os.makedirs(DATA_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(SCHEMA); conn.commit()
    cols = ["job_title","company","location","city_normalized","role_category",
            "experience","exp_min","exp_max","salary","skills_extracted","skills_raw",
            "is_fresher_role","scrape_date","scrape_week","job_url","search_query",
            "search_city","platform"]
    cols = [c for c in cols if c in df.columns]
    cur = conn.cursor(); inserted = 0
    for _, row in df[cols].iterrows():
        try:
            cur.execute(f"INSERT OR IGNORE INTO jobs ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})",
                        [row.get(c) for c in cols])
            inserted += cur.rowcount
        except: continue
    conn.commit()
    week = df["scrape_week"].iloc[0]
    aggregate(conn, week)
    conn.close()
    print(f"  ✓ {inserted} new rows inserted | Week: {week}")
    return inserted

def save_xlsx(df):
    """Save styled Excel with 3 sheets"""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from collections import Counter

    os.makedirs(DATA_DIR, exist_ok=True)

    # Sheet 2: platform breakdown
    platform_df = df.groupby(["platform","role_category"]).size().reset_index(name="Job Count")

    # Sheet 3: skill frequency
    all_skills = []
    for s in df["skills_extracted"].dropna():
        if s not in ("N/A",""):
            all_skills.extend([x.strip() for x in s.split(",")])
    skill_df = pd.DataFrame(Counter(all_skills).most_common(50), columns=["Skill","Frequency"])
    skill_df["% of Jobs"] = (skill_df["Frequency"]/max(len(df),1)*100).round(1).astype(str)+"%"

    display_cols = ["job_title","company","platform","location","city_normalized",
                    "role_category","experience","salary","skills_extracted","job_url","scrape_date"]
    display_cols = [c for c in display_cols if c in df.columns]

    with pd.ExcelWriter(XLSX_OUTPUT, engine="openpyxl") as w:
        df[display_cols].to_excel(w, sheet_name="All Jobs", index=False)
        platform_df.to_excel(w, sheet_name="Platform Breakdown", index=False)
        skill_df.to_excel(w, sheet_name="Skill Frequency", index=False)

    wb = load_workbook(XLSX_OUTPUT)
    colors = {"All Jobs":"1F4E79","Platform Breakdown":"375623","Skill Frequency":"7B3F00"}
    for sheet, color in colors.items():
        ws = wb[sheet]
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len+4, 60)
        ws.freeze_panes = "A2"
    wb.save(XLSX_OUTPUT)
    print(f"  ✓ Excel saved → {XLSX_OUTPUT}")

def run():
    if not os.path.exists(CSV_OUTPUT):
        print(f"No CSV at {CSV_OUTPUT}. Run scraper first."); return
    df = pd.read_csv(CSV_OUTPUT)
    print(f"Raw rows: {len(df)}")
    df = filter_locations(df)
    df.drop_duplicates(subset=["URL","Title","Company"], inplace=True)

    now = datetime.now()
    df["job_title"]        = df.get("Title","")
    df["company"]          = df.get("Company","")
    df["location"]         = df.get("Location","")
    df["experience"]       = df.get("Experience","N/A")
    df["salary"]           = df.get("Salary","N/A")
    df["skills_raw"]       = df.get("Skills","N/A")
    df["job_url"]          = df.get("URL","")
    df["search_query"]     = df.get("Role Searched","")
    df["platform"]         = df.get("Platform","")
    df["city_normalized"]  = df["location"].apply(normalize_city)
    df["role_category"]    = df["job_title"].apply(categorize)
    df["skills_extracted"] = (df["skills_raw"].fillna("")+" "+df["job_title"].fillna("")).apply(extract_skills)
    df[["exp_min","exp_max"]] = df["experience"].apply(lambda x: pd.Series(parse_exp(x)))
    df["is_fresher_role"]  = df["exp_min"].apply(lambda x: 1 if x==0 else 0)
    df["salary_min_lpa"]   = None
    df["salary_max_lpa"]   = None
    df["scrape_date"]      = now.strftime("%Y-%m-%d")
    df["scrape_week"]      = now.strftime("%Y-W%V")
    df["search_city"]      = df["city_normalized"]

    load_db(df)
    save_xlsx(df)
    print(f"✅ Done. {len(df)} jobs → DB + Excel")

if __name__ == "__main__":
    run()